import requests
import json
import arrow
import datetime
import csv
import os
import shutil
import string


import pdb


# Shipstation API creds
api_key = '1faba07e6a5845e7a52111ccda1dae23'
api_secret = 'e5e6bdfc7333440a867d2b34ef133b18'
base_endpoint = 'https://ssapi.shipstation.com/'

# Master list of coffee sizes
coffee_sizes = ['10oz', '1lb', 'KILO', '3lb', '5lb']

# Filename for the wholesale spreadsheet
filename = 'Wholesale Pounds.xlsx'

def open_workbook():
	from openpyxl import load_workbook
	wb = load_workbook(filename=filename)
	return wb

def close_workbook(wb):
	wb.save(filename)

def make_api_query(endpoint, params):
	api_endpoint = endpoint + '?' + params
	url = '{}{}'.format(base_endpoint, api_endpoint)
	r = requests.get(url, auth=(api_key, api_secret))
	results = r.json()
	return results	

def get_wholesale_customers():
	endpoint = 'customers'
	params = 'tagId=49965'
	results = make_api_query(endpoint, params)
	return results.get('customers')

def most_recent_week():
	# Open the spreadsheet
	wb = open_workbook()
	sheet = wb.get_active_sheet()

	last_week_num = None
	for col in sheet.iter_cols(min_row=1, max_row=1):

		if col[0].value == "Company":
			# Skip the company column
			continue
		elif "Week" in col[0].value:
			# Mark the last week number
			last_week_num = int(col[0].value.split(" ")[1])

	return last_week_num

def current_week_num():
	return datetime.datetime.now().isocalendar()[1]

def week_num_to_dates(last_week_num):
	# Convert the week number to special string format
	week_num = "%d-W%d" % (2017, last_week_num)

	# Get the start date, and subtract 3 days to make it the previous Thursday at midnight
	start_date = (datetime.datetime.strptime(week_num + '-0', "%Y-W%W-%w") + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)

	# Get the end date by adding 6 days to the start date, making it Wednesday at 11:59pm
	end_date = (start_date + datetime.timedelta(days=6)).replace(hour=23, minute=59, second=59, microsecond=0)

	return str(start_date), str(end_date)

def num_to_letter(num):
	""" Convert the cell number to a column letter
	"""
	title = ''
	alist = string.uppercase
	while num:
		mod = (num-1) % 26
		num = int((num - mod) / 26)  
		title += alist[mod]

	ret = title[::-1]

	return ret

def make_new_week_column(week_num):
	""" Finds the last column and creates a new column header for the next week number
	"""
	# Open the spreadsheet
	wb = open_workbook()
	sheet = wb.get_active_sheet()

	col_idx = 0
	for col in sheet.iter_cols(min_row=1, max_row=1):

		if col[0].value == "Company" or "Week" in col[0].value:
			# Skip the company column
			# Continue on to the next column, since this week has already been done
			col_idx += 1
			continue

	# Convert the column index to an excel spreadsheet column letter
	col_idx += 1  # Add 1 to the number since the indexing starts with 0
	col_letter = num_to_letter(col_idx)
	col_coord = "%s1" % col_letter

	# Add the cell to the sheet
	sheet[col_coord] = "Week %d" % week_num

	close_workbook(wb)

	return col_idx, col_letter

def convert_fulfillment_to_shipment(fulfillments):
	
	fulfillments["shipments"] = fulfillments["fulfillments"]

	for fulfillment in fulfillments.get('shipments'):
		# Query to get the order
		endpoint = 'orders'
		params = 'orderNumber=%s' % fulfillment.get('orderNumber')
		order = make_api_query(endpoint, params)
		
		order = order.get('orders')[0]
		fulfillment["shipmentItems"] = order["items"]

	return fulfillments

def combine_shipments_and_fulfillments(shipment_results, fulfillment_results):
	fulfillments = convert_fulfillment_to_shipment(fulfillment_results)
	shipment_results['shipments'] += fulfillments.get('shipments')
	return shipment_results

def shipments_by_customer(customer, start_date, end_date):
	# Query for shipments by customer for the current week
	endpoint = 'shipments'
	params = 'shipDateStart=' + start_date +  '&shipDateEnd=' + end_date + '&recipientName=' + customer.get('name') + '&includeShipmentItems=true'
	shipment_results = make_api_query(endpoint, params)

	# Query for fulfillments by customer for the current week
	endpoint = 'fulfillments'
	params = 'shipDateStart=' + start_date +  '&shipDateEnd=' + end_date + '&recipientName=' + customer.get('name')
	fulfillment_results = make_api_query(endpoint, params)

	# Combine shipments and fulfillments
	if shipment_results.get('total') > 0 and fulfillment_results.get('total') > 0:
		results = combine_shipments_and_fulfillments(shipment_results, fulfillment_results)
	elif shipment_results.get('total') > 0 and fulfillment_results.get('total') == 0:
		results = shipment_results
	elif fulfillment_results.get('total') > 0 and shipment_results.get('total') == 0:
		results = convert_fulfillment_to_shipment(fulfillment_results)
	else:
		results = shipment_results

	return results

def coffee_sizes_in_shipments(shipments):
	coffee_dict = {}

	def item_to_dict(item):
		for curr_size in coffee_sizes:
			if curr_size in item.get('sku'):
				if not coffee_dict.get(curr_size):
					coffee_dict[curr_size] = 0
				coffee_dict[curr_size] += item.get('quantity')

	for shipment in shipments.get('shipments'):
		for item in shipment.get('shipmentItems'):
			if 'CFE' in item.get('sku'):
				item_to_dict(item)
	
	return coffee_dict

def coffee_sizes_to_pounds(coffee_size_dict):
	total_pounds = 0.0
	for size, quantity in coffee_size_dict.iteritems():
		if size == '10oz':
			total_pounds += (quantity*10/16)
		elif size == '1lb':
			total_pounds += quantity
		elif size == 'KILO':
			total_pounds += (quantity*2.2)
		elif size == '3lb':
			total_pounds += (quantity*3)
		elif size == '5lb':
			total_pounds += (quantity*5)
	return total_pounds

def populate_spreadsheet(customer, customer_pounds, col_num, col_letter):
	# Open the spreadsheet
	wb = open_workbook()
	sheet = wb.get_active_sheet()

	for row in sheet.iter_rows(min_row=2, min_col=1, max_col=col_num):
		if row[0].value == customer.get('company'):
			sheet[row[-1].coordinate] = customer_pounds

	close_workbook(wb)

# Get a list of the wholesale customers
customers = get_wholesale_customers()

most_recent_week_num = most_recent_week()
curr_week_num = current_week_num()

start_week = most_recent_week_num + 1

# Loop through all weeks until this week
while (curr_week_num > start_week):
	# Get start/end date for week number
	start_date, end_date = week_num_to_dates(start_week)

	# Write new week to spreadsheet column
	col_num, col_letter = make_new_week_column(start_week)

	for customer in customers:
		# Get the shipments for the customer in the specified week
		shipments = shipments_by_customer(customer, start_date, end_date)
		
		# Get a dict of quantities of all coffee sizes in shipment
		coffee_size_dict = coffee_sizes_in_shipments(shipments)
		
		# Convert coffee size quantities to a single total pounds number
		customer_pounds = coffee_sizes_to_pounds(coffee_size_dict)

		if customer_pounds > 0:
			# TO DO - Populate spreadsheet with customer pounds
			populate_spreadsheet(customer, customer_pounds, col_num, col_letter)

			print "%s - Week # %d - %.2f" % (customer.get('name'), start_week, customer_pounds)

	print "Wrote week %d" % start_week

	start_week += 1
